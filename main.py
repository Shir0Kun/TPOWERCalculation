from fastapi import FastAPI, UploadFile, File, Response
import io
import json
import traceback
import pandas as pd
from openpyxl import load_workbook
from collections import Counter
from typing import Dict, List
import tempfile
import os
from datetime import datetime
import re

app = FastAPI()

def process_excel_data(contents):
    """处理Excel数据的通用函数"""
    excel_bytes = io.BytesIO(contents)
    wb = load_workbook(excel_bytes, data_only=True)
    sheet = wb.active

    # 读取所有数据
    max_row = sheet.max_row
    max_col = sheet.max_column
    all_data = []
    
    for row_idx in range(1, max_row + 1):
        row_data = []
        for col_idx in range(1, max_col + 1):
            cell_value = sheet.cell(row=row_idx, column=col_idx).value
            row_data.append(cell_value)
        
        if any(cell is not None and str(cell).strip() != "" for cell in row_data):
            all_data.append(row_data)
    
    if not all_data:
        return None, "Excel文件为空"

    # 检测标题行
    header_row_index = None
    for i, row in enumerate(all_data):
        row_text = " ".join(str(cell) for cell in row if cell)
        if any(keyword in row_text for keyword in ["代理", "ID", "金额", "BONUS", "上分"]):
            header_row_index = i
            break

    if header_row_index is None:
        header_row_index = 0

    # 创建DataFrame
    headers = [str(c).strip() if c else f"Column_{i}" for i, c in enumerate(all_data[header_row_index])]
    data_rows = all_data[header_row_index + 1:]
    
    df = pd.DataFrame(data_rows, columns=headers)
    df = df.dropna(how="all")
    df = df.reset_index(drop=True)
    df = df.where(pd.notnull(df), None)
    
    return df, None

def extract_agent_number(agent_str):
    """从代理序号中提取数字，处理各种格式"""
    if not agent_str:
        return 0
    
    agent_str = str(agent_str).strip()
    
    # 尝试提取数字
    numbers = re.findall(r'\d+', agent_str)
    if numbers:
        return int(numbers[0])
    else:
        return 0

def sort_by_agent_number(df):
    """按代理序号排序：先按上分地方，再按代理序号"""
    if '上分地方' in df.columns and '代理序号' in df.columns:
        # 创建排序键：先按地点，再按代理序号类型，最后按数字
        def sort_key(row):
            location = str(row['上分地方']) if row['上分地方'] is not None else ""
            agent = str(row['代理序号']) if row['代理序号'] is not None else ""
            
            # 地点排序
            location_order = {
                "BELLA": 1,
                "肥子代理 638": 2,
                "OC619-01-01": 3,
                "WS": 4,
                "OC619-01-01-01": 5
            }
            
            # 代理序号排序
            agent_upper = agent.upper()
            if "IN" in agent_upper:
                agent_type = 1  # IN类型在前
            elif "OUT" in agent_upper:
                agent_type = 2  # OUT类型在后
            else:
                agent_type = 3  # 其他类型在最后
            
            # 提取数字
            agent_number = extract_agent_number(agent)
            
            return (location_order.get(location, 99), agent_type, agent_number, location, agent)
        
        # 应用排序
        df_sorted = df.copy()
        df_sorted['sort_key'] = df_sorted.apply(sort_key, axis=1)
        df_sorted = df_sorted.sort_values('sort_key')
        df_sorted = df_sorted.drop('sort_key', axis=1)
        return df_sorted.reset_index(drop=True)
    else:
        # 如果缺少列，就按原始顺序
        return df

def calculate_hierarchical_commission(df):
    """计算层级佣金分配（修正版：每层包含自己业绩）"""
    # 定义层级关系
    hierarchy = {
        # 第一层
        "OC619": {
            "rate": 0.05,  # 5%
            "level": 1,
            "children": ["OC619-01", "OC619-01-01", "OC619-01-02", "OC619-01-03", "OC619-01-01-01"]
        },
        # 第二层
        "OC619-01": {
            "rate": 0.20,  # 20%
            "level": 2,
            "children": ["OC619-01-01", "OC619-01-02", "OC619-01-03", "OC619-01-01-01"]
        },
        # 第三层
        "OC619-01-01": {"rate": 0.05, "level": 3},  # 5%
        "OC619-01-02": {"rate": 0.05, "level": 3},  # 5%
        "OC619-01-03": {"rate": 0.05, "level": 3},  # 5%
        "OC619-01-01-01": {"rate": 0.05, "level": 3},  # 5%
        
        # 其他代理（固定30%）
        "肥子代理 638": {"rate": 0.30, "level": "其他"},
        "BELLA": {"rate": 0.30, "level": "其他"},
        "WS": {"rate": 0.30, "level": "其他"}
    }
    
    # 计算每个地点的总金额
    location_totals = {}
    for location in df['上分地方'].dropna().unique():
        if location:
            location_data = df[df['上分地方'] == location]
            location_totals[location] = location_data['金额'].sum()
    
    # 计算层级佣金
    commission_results = {}
    
    # 先计算每个层级的业绩总和
    level_totals = {
        1: 0,  # 第一层总业绩
        2: 0,  # 第二层总业绩  
        3: 0   # 第三层总业绩
    }
    
    # 确定每个地点属于哪个层级
    location_levels = {}
    for location in location_totals.keys():
        location_str = str(location)
        level = None
        
        if "OC619-01-01-01" in location_str:
            level = 3
        elif "OC619-01-01" in location_str or "OC619-01-02" in location_str or "OC619-01-03" in location_str:
            level = 3
        elif "OC619-01" in location_str:
            level = 2
        elif "OC619" in location_str:
            level = 1
        else:
            level = "其他"
        
        location_levels[location] = level
        if level in [1, 2, 3]:
            level_totals[level] += location_totals[location]
    
    # 计算各层总业绩（包含下层）
    level_cumulative_totals = {
        1: level_totals[1] + level_totals[2] + level_totals[3],  # 第一层+第二层+第三层
        2: level_totals[2] + level_totals[3],  # 第二层+第三层
        3: level_totals[3]  # 第三层
    }
    
    # 为每个地点计算佣金
    for location, total_amount in location_totals.items():
        location_str = str(location)
        level = location_levels[location]
        
        if level in [1, 2, 3]:
            # OC619系列层级计算
            commission_breakdown = {}
            
            if level == 1:
                # 第一层：第一层+第二层+第三层总业绩 × 5%
                commission_breakdown["OC619"] = level_cumulative_totals[1] * hierarchy["OC619"]["rate"]
            
            elif level == 2:
                # 第二层：第二层+第三层总业绩 × 20%
                commission_breakdown["OC619-01"] = level_cumulative_totals[2] * hierarchy["OC619-01"]["rate"]
            
            elif level == 3:
                # 第三层：第三层总业绩 × 5%
                # 确定具体的第三层代理名称
                third_level_agent = None
                for agent in ["OC619-01-01-01", "OC619-01-01", "OC619-01-02", "OC619-01-03"]:
                    if agent in location_str:
                        third_level_agent = agent
                        break
                
                if third_level_agent:
                    commission_breakdown[third_level_agent] = level_cumulative_totals[3] * hierarchy[third_level_agent]["rate"]
                else:
                    commission_breakdown["第三层代理"] = level_cumulative_totals[3] * 0.05
            
            commission_results[location] = {
                "总金额": total_amount,
                "层级": f"第{level}层",
                "佣金分配": commission_breakdown,
                "计算基础": {
                    "第一层业绩": level_totals[1],
                    "第二层业绩": level_totals[2],
                    "第三层业绩": level_totals[3],
                    "第一层计算基础": level_cumulative_totals[1],
                    "第二层计算基础": level_cumulative_totals[2],
                    "第三层计算基础": level_cumulative_totals[3]
                }
            }
        
        else:
            # 非OC619系列：固定30%
            agent_key = None
            for key in ["肥子代理 638", "BELLA", "WS"]:
                if key in location_str:
                    agent_key = key
                    break
            
            if agent_key is None:
                agent_key = location_str
            
            commission_breakdown = {
                agent_key: total_amount * 0.30
            }
            
            commission_results[location] = {
                "总金额": total_amount,
                "层级": "其他代理",
                "佣金分配": commission_breakdown,
                "计算基础": {"自己业绩": total_amount}
            }
    
    return commission_results

@app.post("/upload-excel/")
async def upload_excel(file: UploadFile = File(...)):
    try:
        contents = await file.read()
        df, error = process_excel_data(contents)
        
        if error:
            return {"error": error}

        # 按上分地方分组统计
        location_stats = {}
        location_details = {}
        
        for index, row in df.iterrows():
            location = str(row.get("上分地方", "") or "").strip()
            if not location:
                location = "未知地点"
            
            # 统计每个地点的交易
            if location not in location_stats:
                location_stats[location] = 0
                location_details[location] = []
            
            location_stats[location] += 1
            
            # 记录详细信息
            transaction_info = {
                "代理序号": row.get("代理序号"),
                "ID": row.get("ID"),
                "金额": row.get("金额"),
                "BONUS": row.get("BONUS"),
                "行号": index + 1
            }
            location_details[location].append(transaction_info)

        # 计算每个地点的金额统计
        location_amounts = {}
        for location, transactions in location_details.items():
            amounts = [t["金额"] for t in transactions if t["金额"] is not None]
            in_transactions = [t for t in transactions if "IN" in str(t["代理序号"]).upper()]
            out_transactions = [t for t in transactions if "OUT" in str(t["代理序号"]).upper()]
            
            if amounts:
                location_amounts[location] = {
                    "总交易笔数": len(transactions),
                    "IN笔数": len(in_transactions),
                    "OUT笔数": len(out_transactions),
                    "总金额": sum(amounts),
                    "平均金额": sum(amounts) / len(amounts),
                    "最大金额": max(amounts),
                    "最小金额": min(amounts)
                }

        # 计算层级佣金
        commission_results = calculate_hierarchical_commission(df)

        # 预览数据（排序后的）
        df_sorted = sort_by_agent_number(df)
        preview_data = []
        for i in range(min(10, len(df_sorted))):
            row_data = {col: df_sorted.iloc[i][col] for col in df_sorted.columns}
            preview_data.append(row_data)

        return {
            "success": True,
            "total_rows": len(df),
            "location_summary": {
                "by_location": location_stats,
                "amount_analysis": location_amounts,
                "total_locations": len(location_stats)
            },
            "commission_calculation": commission_results,
            "preview": preview_data,
            "export_ready": True,
            "sorting_applied": "按上分地方和代理序号排序"
        }

    except Exception as e:
        return {"error": str(e), "traceback": traceback.format_exc()}

@app.post("/export-sorted/")
async def export_sorted(file: UploadFile = File(...)):
    """完全排序导出 - 先按上分地方，再按代理序号，包含修正的层级佣金计算"""
    try:
        contents = await file.read()
        df, error = process_excel_data(contents)
        
        if error:
            return {"error": error}

        # 完全排序：先按上分地方，再按代理序号
        df_sorted = sort_by_agent_number(df)
        
        # 计算层级佣金
        commission_results = calculate_hierarchical_commission(df)
        
        # 准备佣金数据
        commission_data = []
        total_commission_by_agent = {}
        
        for location, result in commission_results.items():
            row_data = {
                '上分地方': location,
                '总金额': result["总金额"],
                '层级': result["层级"]
            }
            
            # 添加计算基础信息
            if "计算基础" in result:
                for key, value in result["计算基础"].items():
                    row_data[f'计算基础_{key}'] = value
            
            # 添加佣金分配
            for agent, commission in result["佣金分配"].items():
                row_data[f'佣金_{agent}'] = commission
                
                # 汇总总佣金
                if agent not in total_commission_by_agent:
                    total_commission_by_agent[agent] = 0
                total_commission_by_agent[agent] += commission
            
            commission_data.append(row_data)

        # 导出
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Sheet 1: 排序后的数据
            df_sorted.to_excel(writer, sheet_name='Sorted_Data', index=False)
            
            # Sheet 2: 地点统计
            location_summary = df_sorted.groupby('上分地方').agg({
                '代理序号': 'count',
                '金额': ['sum', 'mean', 'min', 'max']
            }).round(2)
            
            # 重命名列
            location_summary.columns = ['交易笔数', '总金额', '平均金额', '最小金额', '最大金额']
            location_summary.to_excel(writer, sheet_name='Location_Statistics')
            
            # Sheet 3: 层级佣金计算
            commission_df = pd.DataFrame(commission_data)
            commission_df.to_excel(writer, sheet_name='Hierarchical_Commission', index=False)
            
            # Sheet 4: 佣金汇总
            summary_data = []
            for agent, total_commission in total_commission_by_agent.items():
                summary_data.append({
                    '代理层级': agent,
                    '总佣金': total_commission
                })
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Commission_Summary', index=False)
            
            # Sheet 5: 层级规则说明
            rules_df = pd.DataFrame({
                '层级佣金规则': [
                    '=== OC619系列层级结构（修正版）===',
                    '',
                    '第一层: OC619',
                    '  - 佣金: (第一层 + 第二层 + 第三层总业绩) × 5%',
                    '  - 计算基础: 包含自己业绩和所有下级业绩',
                    '',
                    '第二层: OC619-01', 
                    '  - 佣金: (第二层 + 第三层总业绩) × 20%',
                    '  - 计算基础: 包含自己业绩和下级第三层业绩',
                    '',
                    '第三层: OC619-01-01/OC619-01-02/OC619-01-03/OC619-01-01-01',
                    '  - 佣金: (全部第三层业绩) × 5%',
                    '  - 计算基础: 所有第三层代理的总业绩',
                    '',
                    '=== 其他代理 ===',
                    '肥子代理 638: 总金额 × 30%',
                    'BELLA: 总金额 × 30%',
                    'WS: 总金额 × 30%',
                    '其他代理: 总金额 × 30%',
                    '',
                    '=== 业绩计算说明 ===',
                    '第一层总业绩: 所有OC619代理的业绩总和',
                    '第二层总业绩: 所有OC619-01代理的业绩总和',
                    '第三层总业绩: 所有OC619-01-01/02/03/01-01代理的业绩总和'
                ]
            })
            rules_df.to_excel(writer, sheet_name='Hierarchy_Rules', index=False)

        output.seek(0)
        
        # 纯英文文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"hierarchical_commission_corrected_{timestamp}.xlsx"
        
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        return {"error": str(e), "traceback": traceback.format_exc()}

@app.get("/")
def root():
    return {
        "message": "Excel Sorting and Corrected Hierarchical Commission API",
        "endpoints": {
            "/upload-excel/": "Preview data with sorting and corrected hierarchical commission",
            "/export-sorted/": "Fully sorted export with corrected hierarchical commission calculation"
        },
        "corrected_hierarchy_rules": {
            "第一层 (OC619)": "佣金: (第一层 + 第二层 + 第三层总业绩) × 5%",
            "第二层 (OC619-01)": "佣金: (第二层 + 第三层总业绩) × 20%", 
            "第三层 (OC619-01-01/02/03/01-01)": "佣金: (全部第三层业绩) × 5%",
            "其他代理": "佣金: (自己业绩) × 30%"
        }
    }

